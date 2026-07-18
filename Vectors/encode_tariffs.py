"""Resumably embed tariff descriptions from trf.xlsx into Chroma."""

import argparse
import json
import random
import time
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from semantic_search.config import Settings, load_settings
from semantic_search.vector_store import SemanticSearchService


DATA_FILE = Path(__file__).parent / "semantic_search" / "data" / "trf.xlsx"


@dataclass(frozen=True)
class TariffRow:
    excel_row: int
    tariff_code: str
    description: str


def read_tariffs(path: Path) -> list[TariffRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iterator = sheet.iter_rows(values_only=True)
        header_row = next(row_iterator)
        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_row)
            if value is not None
        }
        try:
            code_column = headers["کد تعرفه"]
            description_column = headers["شرح"]
        except KeyError as error:
            raise ValueError(
                "ستون‌های «کد تعرفه» و «شرح» در فایل پیدا نشدند."
            ) from error

        rows: list[TariffRow] = []
        for excel_row, values in enumerate(row_iterator, start=2):
            code_value = values[code_column]
            description_value = values[description_column]
            tariff_code = str(code_value).strip() if code_value is not None else ""
            description = (
                str(description_value).strip()
                if description_value is not None
                else ""
            )
            if tariff_code and description:
                rows.append(TariffRow(excel_row, tariff_code, description))
        return rows
    finally:
        workbook.close()


def load_next_row(checkpoint: Path) -> int:
    if not checkpoint.exists():
        return 2
    try:
        data = json.loads(checkpoint.read_text(encoding="utf-8"))
        return max(2, int(data["next_excel_row"]))
    except (KeyError, TypeError, ValueError, json.JSONDecodeError):
        raise RuntimeError(f"Checkpoint is invalid: {checkpoint}") from None


def save_checkpoint(checkpoint: Path, next_excel_row: int) -> None:
    checkpoint.parent.mkdir(parents=True, exist_ok=True)
    temporary = checkpoint.with_suffix(".tmp")
    temporary.write_text(
        json.dumps({"next_excel_row": next_excel_row}, indent=2),
        encoding="utf-8",
    )
    temporary.replace(checkpoint)


def reset(settings: Settings) -> SemanticSearchService:
    service = SemanticSearchService(settings)
    try:
        service.reset_collection()
    except ValueError:
        pass
    settings.encoding_checkpoint.unlink(missing_ok=True)
    return SemanticSearchService(settings)


def encode(
    *,
    settings: Settings,
    rows: list[TariffRow],
    batch_size: int,
    delay: float,
    retries: int,
    reset_collection: bool,
) -> None:
    service = reset(settings) if reset_collection else SemanticSearchService(settings)
    next_row = load_next_row(settings.encoding_checkpoint)
    pending = [row for row in rows if row.excel_row >= next_row]
    completed = len(rows) - len(pending)

    print(f"Loaded {len(rows):,} tariff rows from {DATA_FILE}")
    print(f"Starting at Excel row {next_row:,} ({completed:,} already completed)")

    for start in range(0, len(pending), batch_size):
        batch = pending[start : start + batch_size]
        payload = [
            (row.excel_row, row.tariff_code, row.description)
            for row in batch
        ]

        for attempt in range(1, retries + 1):
            try:
                service.add_tariffs(payload)
                break
            except Exception as error:
                if "403 Forbidden" in str(error):
                    print(
                        "\nGemini denied access (HTTP 403). Check the API key, "
                        "project access, and regional availability."
                    )
                    raise
                if attempt == retries:
                    print(f"\nFailed at Excel row {batch[0].excel_row}: {error}")
                    print("Run the same command again to resume from this batch.")
                    raise
                backoff = delay * (2 ** (attempt - 1)) + random.uniform(0, 1)
                print(
                    f"\nAttempt {attempt}/{retries} failed: {error}\n"
                    f"Retrying in {backoff:.1f} seconds..."
                )
                time.sleep(backoff)

        next_excel_row = batch[-1].excel_row + 1
        save_checkpoint(settings.encoding_checkpoint, next_excel_row)
        completed += len(batch)
        percent = completed / len(rows) * 100
        print(
            f"\rEncoded {completed:,}/{len(rows):,} rows "
            f"({percent:5.1f}%) — checkpoint: Excel row {next_excel_row:,}",
            end="",
            flush=True,
        )
        if start + batch_size < len(pending) and delay > 0:
            time.sleep(delay)

    print("\nEncoding complete. All vectors are persisted in Chroma.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Embed trf.xlsx descriptions into the local Chroma database."
    )
    parser.add_argument("--reset", action="store_true", help="delete existing vectors first")
    parser.add_argument("--batch-size", type=int, default=20, choices=range(1, 101))
    parser.add_argument(
        "--delay",
        type=float,
        default=13.0,
        help="seconds between batches (13s keeps 20-row batches under 100 items/minute)",
    )
    parser.add_argument("--retries", type=int, default=5)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.delay < 0 or args.retries < 1:
        raise ValueError("delay must be >= 0 and retries must be >= 1")
    encode(
        settings=load_settings(),
        rows=read_tariffs(DATA_FILE),
        batch_size=args.batch_size,
        delay=args.delay,
        retries=args.retries,
        reset_collection=args.reset,
    )


if __name__ == "__main__":
    main()
